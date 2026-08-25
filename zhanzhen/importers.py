"""鼎信诺/金蝶 账套文件导入器（xlsx/CSV → VoucherDraft 草稿）。

职责边界（诚实原则，spec §4.2）：
- 本模块只做「格式识别 + 列模糊匹配 + 行→草稿」的结构化提取，不编造任何数据；
- 每行凭证由 service 层走完整管线（ingest→OCR→分录草稿→确认门），不绕过
  哈希链与状态机；
- openpyxl 是可选依赖（pip install 'zhanzhen[excel]'）：未安装时 xlsx 一律报
  unknown 并给出安装提示，绝不静默降级。

格式识别规则（detect_format）：
- 数据以 PK\\x03\\x04 开头 → xlsx：读各 sheet 名与首行列头，
  含「凭证明细|对方单位」→ dingxinuo；含「科目编码」→ kingdee；否则 unknown；
- 非 xlsx → 按 CSV 处理（utf-8-sig/utf-8/gbk 依次尝试解码），首行列头命中
  ≥2 个记账关键词 → generic_csv；否则 unknown。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

__all__ = ["VoucherDraft", "detect_format", "import_dingxinuo", "pickCol"]

_XLSX_MAGIC = b"PK\x03\x04"

# 鼎信诺凭证明细特征：sheet 名或首行列头出现任一关键词即可判定
_DINGXINUO_KEYWORDS = ("凭证明细", "对方单位")
# 金蝶导出特征
_KINGDEE_KEYWORDS = ("科目编码",)

# CSV 列头判定用的记账关键词（命中 ≥2 个才算通用账套 CSV）
_CSV_KEYWORDS = ("日期", "凭证", "摘要", "科目", "借方", "贷方", "对方")

# 模糊列匹配：字段名 → 关键词优先级列表（同一字段先命中的关键词优先）
_FIELD_KEYWORDS: dict = {
    "date": ("记账日期", "凭证日期", "单据日期", "日期"),
    "voucher_no": ("凭证编号", "凭证号", "单据编号"),
    "summary": ("摘要",),
    "account": ("科目名称", "科目全名", "科目编码", "科目"),
    "debit": ("借方金额", "借方发生额", "借方", ),
    "credit": ("贷方金额", "贷方发生额", "贷方", ),
    "counterparty": ("对方单位", "对方名称", "往来单位", "客户名称", "供应商名称"),
}

# 明细表表头至少要能认出这些字段之一，否则不当作明细 sheet
_HEADER_HINTS = _CSV_KEYWORDS


@dataclass
class VoucherDraft:
    """一行账套分录的中性草稿表示——不绑定具体软件格式。"""

    date: str = ""
    voucher_no: str = ""
    summary: str = ""
    account: str = ""
    debit: float = 0.0
    credit: float = 0.0
    counterparty: str = ""


def _cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "none" else s


def pickCol(headers, keywords) -> int | None:
    """模糊列匹配：返回第一个包含任一关键词的列下标；找不到返回 None。

    keywords 有优先级：先按关键词顺序找整行，再退到下一个关键词。
    """
    norm = [_cell_str(h).replace(" ", "").replace("\u3000", "") for h in headers]
    for kw in keywords:
        for i, h in enumerate(norm):
            if h and kw in h:
                return i
    return None


def detect_format(data: bytes) -> str:
    """识别账套文件格式：dingxinuo | kingdee | generic_csv | unknown。"""
    if not data:
        return "unknown"
    if data[:4] == _XLSX_MAGIC:
        try:
            import openpyxl  # 可选依赖，懒加载
        except ImportError:
            import warnings
            warnings.warn(
                "检测到 xlsx 文件但未安装 openpyxl："
                "请执行 pip install 'zhanzhen[excel]' 后重试")
            return "unknown"
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            return "unknown"
        try:
            blob_parts = []
            for ws in wb.worksheets:
                blob_parts.append(_cell_str(ws.title))
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    blob_parts.extend(_cell_str(c) for c in row)
                    break
            text = "\n".join(blob_parts)
        finally:
            wb.close()
        if any(kw in text for kw in _DINGXINUO_KEYWORDS):
            return "dingxinuo"
        if any(kw in text for kw in _KINGDEE_KEYWORDS):
            return "kingdee"
        return "unknown"
    # 非 xlsx：尝试 utf-8-sig / utf-8 / gbk 解码为 CSV 判断
    text = None
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = data.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        return "unknown"
    first_line = next((ln for ln in text.splitlines() if ln.strip()), "")
    cells = [c.strip().strip('"').replace(" ", "")
             for c in re.split(r"[,\t;]", first_line.strip())]
    hits = sum(1 for c in cells if any(kw in c for kw in _CSV_KEYWORDS))
    return "generic_csv" if hits >= 2 else "unknown"


_DATE_RE = re.compile(r"(20\d{2})[-年/.](\d{1,2})[-月/.](\d{1,2})")
_MONEY_CLEAN_RE = re.compile(r"[,¥￥\s]")


def _norm_date(v) -> str:
    """Excel 日期单元格(datetime)或字符串 → ISO 'YYYY-MM-DD'；解析不了原样返回。"""
    import datetime as _dt
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    s = _cell_str(v)
    if not s:
        return ""
    m = _DATE_RE.search(s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s


def _norm_money(v) -> float:
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v) + 1e-9, 2)
    s = _MONEY_CLEAN_RE.sub("", str(v))
    try:
        return round(float(s) + 1e-9, 2)
    except ValueError:
        return 0.0


_TOTAL_MARKS = ("合计", "总计", "小计", "累计")


def _row_to_draft(row, col: dict) -> VoucherDraft | None:
    """单行 → VoucherDraft；全空行与合计/小计行返回 None（跳过）。"""
    cells = list(row)
    width = max((i for i in col.values() if i is not None), default=-1) + 1
    if len(cells) < width:
        cells += [None] * (width - len(cells))
    texts = [_cell_str(c) for c in cells]
    if not any(texts):                      # 全空行
        return None
    joined = "".join(texts)
    if any(mk in joined for mk in _TOTAL_MARKS):   # 合计/小计行
        return None

    def get(field):
        i = col.get(field)
        return cells[i] if i is not None and i < len(cells) else None

    no = _cell_str(get("voucher_no"))
    if no.endswith(".0"):                   # Excel 把纯数字单号读成 float
        no = no[:-2]
    draft = VoucherDraft(
        date=_norm_date(get("date")),
        voucher_no=no,
        summary=_cell_str(get("summary")),
        account=_cell_str(get("account")),
        debit=_norm_money(get("debit")),
        credit=_norm_money(get("credit")),
        counterparty=_cell_str(get("counterparty")),
    )
    return draft


def _looks_like_detail_header(header_row) -> bool:
    text = "".join(_cell_str(c) for c in header_row)
    return sum(1 for kw in _HEADER_HINTS if kw in text) >= 2


def import_dingxinuo(data: bytes) -> list[VoucherDraft]:
    """解析鼎信诺「凭证明细」xlsx → 草稿列表。

    - 自动定位含明细表头的 sheet（支持多 sheet）；
    - 模糊列匹配（pickCol）容忍列序与列名差异；
    - 跳过全空行与合计行。
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "导入鼎信诺 xlsx 需要 openpyxl：pip install 'zhanzhen[excel]'") from e

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    drafts: list[VoucherDraft] = []
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = None
            for row in rows:
                texts = [_cell_str(c) for c in row]
                if any(texts):
                    header = texts
                    break
            if not header or not _looks_like_detail_header(header):
                continue                        # 非明细 sheet，跳过
            col = {f: pickCol(header, kws) for f, kws in _FIELD_KEYWORDS.items()}
            has_amount_col = col.get("debit") is not None or col.get("credit") is not None
            if col.get("date") is None and not has_amount_col:
                continue                        # 认不出日期也认不出金额，不当明细处理
            for row in rows:
                d = _row_to_draft(row, col)
                if d is not None:
                    drafts.append(d)
    finally:
        wb.close()
    return drafts
